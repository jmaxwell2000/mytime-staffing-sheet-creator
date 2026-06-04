"""Build print-ready staffing workbooks."""

from __future__ import annotations

from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, time
from io import BytesIO
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.parser.constants import EXCLUDED_JOB_CODES, ROLE_OTHER, ROLE_PSA, ROLE_RN, SHIFT_DAY, SHIFT_NIGHT
from app.parser.models import ScheduleRecord

DAY_PERIODS = ("0700-1100", "1100-1500", "1500-1900")
NIGHT_PERIODS = ("1900-2300", "2300-0300", "0300-0700")

RN_CAPACITY = 15
PSA_CAPACITY = 11
BACK_LINE_ROWS = tuple(range(3, 30))
COLUMN_WIDTHS = {"A": 2, "B": 4.14, "C": 17.5, "D": 5.5, "E": 5.5, "F": 22.95, "G": 22.95, "H": 22.95}
FRONT_PENCIL_REMINDER = "Use pencil. Write and initial notes on the back of this sheet."
BACK_PENCIL_REMINDER = "Use pencil"

HEADER_FILL = PatternFill("solid", fgColor="D9E3F0")
UNORTHODOX_TIME_FILL = PatternFill("solid", fgColor="FFF2CC")
FALL_TITLE_FILL = PatternFill("solid", fgColor="FCE4D6")
FALL_HEADER_FILL = PatternFill("solid", fgColor="FBE5D6")
FALL_BODY_FILL = PatternFill("solid", fgColor="FFF8E1")
THIN = Side(style="thin", color="000000")
BACK_LINE_SIDE = Side(style="thin", color="B7B7B7")
BACK_TEXT_COLOR = "808080"
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
QUESTION_BULLET = "\u2022 "
FALL_PREVENTION_TEXT = "Post-fall measures (circle):              Alarm           Nonskid socks             Fall sign             Stay with pt toileting             Fall risk score             Other:____________"
STAFF_NAME_BASE_FONT_SIZE = 11
STAFF_NAME_MIN_FONT_SIZE = 7
STAFF_NAME_FIT_CHARS_AT_BASE_SIZE = 19


@dataclass(frozen=True)
class SheetLayout:
    total_rows: int
    back_line_rows: tuple[int, ...]


def build_staffing_workbook(
    records: Iterable[ScheduleRecord],
    *,
    include_back_pages: bool = False,
    back_page_offset: float = 0.0,
) -> bytes:
    record_list = [record for record in records if record.job_code not in EXCLUDED_JOB_CODES]
    workbook = Workbook()
    workbook.remove(workbook.active)

    grouped = _group_records(record_list)
    for work_date in sorted(grouped.keys(), reverse=True):
        for shift_kind in (SHIFT_NIGHT, SHIFT_DAY):
            shift_records = grouped[work_date].get(shift_kind)
            if not shift_records:
                continue
            sheet_name = _sheet_name(work_date, shift_kind)
            front = workbook.create_sheet(sheet_name)
            layout = _build_front_sheet(front, work_date, shift_kind, shift_records)
            if include_back_pages:
                back = workbook.create_sheet(f"{sheet_name} BACK")
                _build_back_sheet(back, work_date, shift_kind, back_page_offset, front, layout)

    if not workbook.worksheets:
        workbook.create_sheet("No Records")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def filename_for_records(records: Iterable[ScheduleRecord]) -> str:
    dates = sorted({record.work_date for record in records})
    if not dates:
        return "staffing-sheets.xlsx"
    if dates[0] == dates[-1]:
        return f"staffing-sheets-{dates[0].isoformat()}.xlsx"
    return f"staffing-sheets-{dates[0].isoformat()}-to-{dates[-1].isoformat()}.xlsx"


def resolve_lead_display(records: Iterable[ScheduleRecord]) -> dict[int, tuple[str, bool]]:
    """Return `{id(record): (display_name, show_lead)}` for a single shift."""

    record_list = list(records)
    leaders = [record for record in record_list if record.is_clinical_leader]
    result: dict[int, tuple[str, bool]] = {}

    if len(leaders) == 1:
        leader_id = id(leaders[0])
        for record in record_list:
            result[id(record)] = (normalize_employee_name(record.employee_display_name), id(record) == leader_id)
        return result

    if len(leaders) > 1:
        leader_ids = {id(record) for record in leaders}
        for record in record_list:
            display_name = normalize_employee_name(record.employee_display_name)
            name = f"{display_name} (CL)" if id(record) in leader_ids else display_name
            result[id(record)] = (name, False)
        return result

    for record in record_list:
        result[id(record)] = (normalize_employee_name(record.employee_display_name), record.has_lead_pay)
    return result


def _build_front_sheet(
    sheet: Worksheet,
    work_date: date,
    shift_kind: str,
    records: list[ScheduleRecord],
) -> SheetLayout:
    _apply_common_page_setup(sheet)
    sheet.sheet_view.showGridLines = False

    _set_front_dimensions(sheet)
    _front_title_row(sheet, work_date, shift_kind)

    periods = NIGHT_PERIODS if shift_kind == SHIFT_NIGHT else DAY_PERIODS
    display = resolve_lead_display(records)

    sections = {
        ROLE_RN: _sort_records([record for record in records if record.role_group == ROLE_RN], display, lead_first=True),
        ROLE_PSA: _sort_records([record for record in records if record.role_group == ROLE_PSA], display),
        ROLE_OTHER: _sort_records([record for record in records if record.role_group == ROLE_OTHER], display),
    }

    rn_header_row = 2
    rn_start_row = rn_header_row + 1
    rn_row_count = _staff_row_count(RN_CAPACITY, sections[ROLE_RN])
    _section_header(sheet, rn_header_row, "RN", periods)
    _staff_rows(sheet, rn_start_row, rn_row_count, sections[ROLE_RN], display, shift_kind)

    psa_header_row = rn_start_row + rn_row_count
    psa_start_row = psa_header_row + 1
    psa_row_count = _staff_row_count(PSA_CAPACITY, sections[ROLE_PSA])
    _section_header(sheet, psa_header_row, "PSA", periods)
    _staff_rows(sheet, psa_start_row, psa_row_count, sections[ROLE_PSA], display, shift_kind)

    fall_start = psa_start_row + psa_row_count
    if sections[ROLE_OTHER]:
        other_header_row = fall_start
        other_start_row = other_header_row + 1
        other_row_count = max(1, len(sections[ROLE_OTHER]))
        _section_header(sheet, other_header_row, "Other", periods)
        _staff_rows(sheet, other_start_row, other_row_count, sections[ROLE_OTHER], display, shift_kind)
        fall_start = other_start_row + other_row_count

    fall_end = _fall_form(sheet, work_date, fall_start)
    spacer_row = fall_end + 1
    bottom_title_row = fall_end + 2
    sheet.row_dimensions[spacer_row].height = 3
    sheet.row_dimensions[bottom_title_row].height = 32
    _bottom_title(sheet, work_date, shift_kind, bottom_title_row)
    sheet.print_area = f"B1:H{bottom_title_row}"
    return SheetLayout(total_rows=bottom_title_row, back_line_rows=tuple(range(3, fall_start)))


def _build_back_sheet(
    sheet: Worksheet,
    work_date: date,
    shift_kind: str,
    offset: float,
    front_sheet: Worksheet,
    layout: SheetLayout,
) -> None:
    effective_top_margin = max(0.05, min(0.45, 0.15 + offset))
    _apply_common_page_setup(sheet)
    sheet.page_margins.left = 0.12
    sheet.page_margins.right = 0.75
    sheet.page_margins.top = effective_top_margin
    sheet.print_area = f"B1:H{layout.total_rows}"
    sheet.sheet_view.showGridLines = False
    _copy_dimensions(front_sheet, sheet, layout.total_rows)

    _back_title_row(sheet, work_date, shift_kind)
    sheet.merge_cells("B3:H3")
    sheet["B3"] = "Notes (Please initial)"
    sheet["B3"].font = Font(bold=True, size=14, color=BACK_TEXT_COLOR)
    sheet["B3"].alignment = Alignment(horizontal="center", vertical="center")

    for row in layout.back_line_rows:
        for column in range(2, 9):
            cell = sheet.cell(row, column)
            cell.border = Border(bottom=BACK_LINE_SIDE)

    _bottom_title(sheet, work_date, shift_kind, layout.total_rows, horizontal="left")
    _apply_back_text_color(sheet, layout.total_rows)


def _apply_back_text_color(sheet: Worksheet, total_rows: int) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=total_rows, min_col=2, max_col=8):
        for cell in row:
            if cell.value is not None:
                font = copy(cell.font)
                font.color = BACK_TEXT_COLOR
                cell.font = font


def _group_records(records: Iterable[ScheduleRecord]) -> dict[date, dict[str, list[ScheduleRecord]]]:
    grouped: dict[date, dict[str, list[ScheduleRecord]]] = defaultdict(lambda: defaultdict(list))
    for record in records:
        grouped[record.work_date][record.shift_kind].append(record)
    return grouped


def _sort_records(
    records: list[ScheduleRecord],
    display: dict[int, tuple[str, bool]] | None = None,
    *,
    lead_first: bool = False,
) -> list[ScheduleRecord]:
    return sorted(
        records,
        key=lambda record: (
            0 if lead_first and display and display[id(record)][1] else 1,
            _last_name_key(display[id(record)][0] if display else normalize_employee_name(record.employee_display_name)),
            record.source_row if record.source_row is not None else 10_000,
        ),
    )


def _staff_row_count(capacity: int, records: list[ScheduleRecord]) -> int:
    return max(capacity, len(records) + 1)


def normalize_employee_name(name: str) -> str:
    compact_name = " ".join(str(name).strip().split())
    if not compact_name:
        return compact_name
    if "," in compact_name:
        last_name, rest = compact_name.split(",", 1)
        return f"{_normalize_name_part(last_name)}, {_normalize_name_part(rest)}"
    return _normalize_name_part(compact_name)


def _normalize_name_part(value: str) -> str:
    compact_value = " ".join(value.strip().split())
    return re.sub(r"[A-Za-z]+", lambda match: _normalize_alpha_word(match.group(0)), compact_value)


def _normalize_alpha_word(value: str) -> str:
    if len(value) == 1:
        return value.upper()
    return f"{value[0].upper()}{value[1:].lower()}"


def _last_name_key(display_name: str) -> tuple[str, str]:
    name = display_name.replace(" (CL)", "")
    if "," in name:
        last_name, rest = name.split(",", 1)
        return (last_name.strip().casefold(), rest.strip().casefold())
    pieces = name.split()
    return (pieces[-1].casefold() if pieces else "", name.casefold())


def _staff_name_font_size(display_name: str) -> int:
    display_length = len(display_name)
    if display_length <= STAFF_NAME_FIT_CHARS_AT_BASE_SIZE:
        return STAFF_NAME_BASE_FONT_SIZE
    fitted_size = int(STAFF_NAME_BASE_FONT_SIZE * STAFF_NAME_FIT_CHARS_AT_BASE_SIZE / display_length)
    return max(STAFF_NAME_MIN_FONT_SIZE, fitted_size)


def _apply_common_page_setup(sheet: Worksheet) -> None:
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.75
    sheet.page_margins.right = 0.12
    sheet.page_margins.top = 0.15
    sheet.page_margins.bottom = 0.15


def _set_front_dimensions(sheet: Worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 32
    for row in range(2, 61):
        sheet.row_dimensions[row].height = 25


def _copy_dimensions(source: Worksheet, target: Worksheet, total_rows: int) -> None:
    for column in range(1, 9):
        letter = get_column_letter(column)
        target.column_dimensions[letter].width = source.column_dimensions[letter].width
    for row in range(1, total_rows + 1):
        target.row_dimensions[row].height = source.row_dimensions[row].height


def _title_row(
    sheet: Worksheet,
    work_date: date,
    shift_kind: str,
    *,
    row: int,
    fill: PatternFill | None,
    font_size: int = 22,
    horizontal: str = "right",
) -> None:
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cell = sheet.cell(row, 2)
    cell.value = _title(work_date, shift_kind)
    cell.font = Font(bold=True, size=font_size)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center")
    if fill:
        for column in range(2, 9):
            sheet.cell(row, column).fill = fill


def _front_title_row(sheet: Worksheet, work_date: date, shift_kind: str) -> None:
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    reminder_cell = sheet.cell(1, 2)
    reminder_cell.value = FRONT_PENCIL_REMINDER
    reminder_cell.font = Font(bold=True, size=9)
    reminder_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True)

    sheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    title_cell = sheet.cell(1, 7)
    title_cell.value = _title(work_date, shift_kind)
    title_cell.font = Font(bold=True, size=22)
    title_cell.alignment = Alignment(horizontal="right", vertical="center", shrink_to_fit=True)


def _back_title_row(sheet: Worksheet, work_date: date, shift_kind: str) -> None:
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    title_cell = sheet.cell(1, 2)
    title_cell.value = _title(work_date, shift_kind)
    title_cell.font = Font(bold=True, size=22)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    reminder_cell = sheet.cell(1, 7)
    reminder_cell.value = BACK_PENCIL_REMINDER
    reminder_cell.font = Font(bold=True, size=10)
    reminder_cell.alignment = Alignment(horizontal="right", vertical="center")


def _bottom_title(
    sheet: Worksheet,
    work_date: date,
    shift_kind: str,
    row: int,
    fill: PatternFill | None = None,
    horizontal: str = "right",
) -> None:
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cell = sheet.cell(row, 2)
    cell.value = _title(work_date, shift_kind)
    cell.font = Font(bold=True, size=22)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center")
    if fill:
        for column in range(2, 9):
            sheet.cell(row, column).fill = fill


def _section_header(sheet: Worksheet, row: int, label: str, periods: tuple[str, str, str]) -> None:
    sheet.row_dimensions[row].height = 30
    values = ["Initial", label, "Start", "End", *periods]
    for offset, value in enumerate(values, start=2):
        cell = sheet.cell(row, offset)
        cell.value = value
        cell.fill = HEADER_FILL
        cell.border = GRID_BORDER
        if offset == 2:
            cell.font = Font(bold=True, size=7)
            cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        else:
            cell.font = Font(bold=True, size=24 if offset == 3 else 11)
            cell.alignment = Alignment(horizontal="center" if offset >= 4 else "left", vertical="center")


def _staff_rows(
    sheet: Worksheet,
    start_row: int,
    row_count: int,
    records: list[ScheduleRecord],
    display: dict[int, tuple[str, bool]],
    shift_kind: str,
) -> None:
    for row_offset in range(row_count):
        row = start_row + row_offset
        for column in range(2, 9):
            cell = sheet.cell(row, column)
            cell.border = GRID_BORDER
            cell.alignment = Alignment(vertical="center")

        if row_offset >= len(records):
            continue

        record = records[row_offset]
        display_name, show_lead = display[id(record)]
        name_cell = sheet.cell(row, 3)
        name_cell.value = display_name
        name_cell.font = Font(size=_staff_name_font_size(display_name))
        name_cell.alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
        _write_time_cell(sheet.cell(row, 4), format_start_time(record.start_time))
        _write_time_cell(sheet.cell(row, 5), format_end_time(record.end_time))
        if show_lead:
            lead_cell = sheet.cell(row, 6)
            lead_cell.value = "LEAD"
            lead_cell.font = Font(bold=True)
            lead_cell.alignment = Alignment(horizontal="center", vertical="center")

        for column, is_standard in (
            (4, is_standard_start(record.start_time, shift_kind)),
            (5, is_standard_end(record.end_time, shift_kind)),
        ):
            if not is_standard:
                cell = sheet.cell(row, column)
                cell.font = Font(size=11)
                cell.fill = UNORTHODOX_TIME_FILL


def _fall_form(sheet: Worksheet, work_date: date, start_row: int) -> int:
    title_row = start_row
    header_row = start_row + 1
    questions_start = start_row + 2
    prevention_row = start_row + 6
    _set_fall_row_heights(sheet, start_row)

    for row in range(title_row, prevention_row + 1):
        for column in range(2, 9):
            cell = sheet.cell(row, column)
            cell.fill = FALL_BODY_FILL
            cell.border = GRID_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    sheet.merge_cells(start_row=title_row, start_column=2, end_row=title_row, end_column=8)
    for column in range(2, 9):
        sheet.cell(title_row, column).fill = FALL_TITLE_FILL
    sheet.cell(title_row, 2).value = (
        f"Patient Fall Review ({_date_short(work_date)}) - Charge RN: if a patient falls, "
        "answer the questions, add brief notes, and report to oncoming Charge RN."
    )
    sheet.cell(title_row, 2).font = Font(bold=True, size=10)
    sheet.cell(title_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=5)
    sheet.cell(header_row, 2).value = "Question"
    sheet.merge_cells(start_row=header_row, start_column=6, end_row=header_row, end_column=8)
    sheet.cell(header_row, 6).value = "Answer / Notes"
    for column in range(2, 9):
        sheet.cell(header_row, column).fill = FALL_HEADER_FILL
    for column in (2, 6):
        sheet.cell(header_row, column).font = Font(bold=True, size=10)
        sheet.cell(header_row, column).alignment = Alignment(horizontal="center", vertical="center")

    questions = [
        "Bathroom or bedside commode?",
        "Staff or family with patient at time of fall?",
        "If no, should someone have been present?",
        "Any injury?",
    ]
    for index, question in enumerate(questions):
        row = questions_start + index
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        sheet.cell(row, 2).value = f"{QUESTION_BULLET}{question}"
        sheet.cell(row, 2).font = Font(size=10)
        sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        sheet.cell(row, 6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    sheet.merge_cells(start_row=prevention_row, start_column=2, end_row=prevention_row, end_column=8)
    sheet.cell(prevention_row, 2).value = FALL_PREVENTION_TEXT
    sheet.cell(prevention_row, 2).font = Font(size=10)
    sheet.cell(prevention_row, 2).alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
    )
    return prevention_row


def _set_fall_row_heights(sheet: Worksheet, start_row: int) -> None:
    row_heights = (26, 24, 30, 30, 30, 30, 30)
    for offset, height in enumerate(row_heights):
        sheet.row_dimensions[start_row + offset].height = height


def format_start_time(value: time) -> str:
    return f"{value.hour:02d}{value.minute:02d}"


def format_end_time(value: time) -> str:
    if value.hour == 19 and value.minute == 0:
        return "1930"
    if value.hour == 7 and value.minute == 0:
        return "0730"
    return f"{value.hour:02d}{value.minute:02d}"


def _write_time_cell(cell, display_value: str) -> None:
    cell.value = int(display_value)
    cell.number_format = "0000"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def is_standard_start(value: time, shift_kind: str) -> bool:
    return format_start_time(value) == ("0700" if shift_kind == SHIFT_DAY else "1900")


def is_standard_end(value: time, shift_kind: str) -> bool:
    return format_end_time(value) == ("1930" if shift_kind == SHIFT_DAY else "0730")


def _sheet_name(work_date: date, shift_kind: str) -> str:
    return f"{work_date.strftime('%a')} {work_date.strftime('%b')}{work_date.day:02d} {shift_kind}"


def _title(work_date: date, shift_kind: str) -> str:
    return f"{work_date.strftime('%a')} {work_date.month}/{work_date.day} - {shift_kind}"


def _date_short(work_date: date) -> str:
    return f"{work_date.strftime('%a')} {work_date.month}/{work_date.day}"
