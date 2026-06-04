"""Evidence-based source workbook parsing."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
import re
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .constants import (
    CLINICAL_LEADER_CODE,
    EXCLUDED_JOB_CODES,
    KNOWN_JOB_CODES,
    LEAD_PAY_CODES,
    PSA_CODES,
    RN_CODES,
    ROLE_OTHER,
    ROLE_PSA,
    ROLE_RN,
    SHIFT_DAY,
    SHIFT_NIGHT,
)
from .models import ParsedSchedule, ScheduleRecord


@dataclass
class HeaderMap:
    kind: str
    job_col: int
    employee_col: int
    start_col: int
    end_col: int
    pay_code_col: int | None = None
    shift_label_col: int | None = None


@dataclass
class PayLeadEvidence:
    employee_display_name: str
    job_code_raw: str
    job_code: str
    start_time: time | None
    end_time: time | None
    source_sheet: str
    source_row: int
    source_evidence: list[str]
    warning: str | None = None


def parse_workbook(data: bytes | BytesIO) -> ParsedSchedule:
    """Parse a MyTime-style source workbook into normalized schedule records."""

    stream = BytesIO(data) if isinstance(data, bytes) else data
    workbook = load_workbook(stream, data_only=True)
    worksheet = workbook["Report"] if "Report" in workbook.sheetnames else workbook.active

    schedule_header: HeaderMap | None = None
    pay_header: HeaderMap | None = None
    records: list[ScheduleRecord] = []
    pay_leads: list[PayLeadEvidence] = []
    warnings: list[str] = []

    for row in worksheet.iter_rows():
        header = _detect_header(row)
        if header:
            if header.kind == "pay":
                pay_header = header
            else:
                schedule_header = header
            continue

        if schedule_header:
            record = _parse_schedule_row(row, schedule_header, worksheet.title)
            if record:
                records.append(record)
                continue

        if pay_header:
            pay_lead = _parse_pay_lead_row(row, pay_header, worksheet.title)
            if pay_lead:
                pay_leads.append(pay_lead)

    _attach_pay_leads(records, pay_leads, warnings)

    if not records:
        warnings.append("No readable schedule records were found.")

    return ParsedSchedule(records=records, warnings=warnings)


def _detect_header(row: Iterable[Any]) -> HeaderMap | None:
    columns: dict[str, int] = {}

    for cell in row:
        text = _normalize_header(cell.value)
        if not text:
            continue
        if text == "job":
            columns.setdefault("job", cell.column)
        elif text == "employee":
            columns.setdefault("employee", cell.column)
        elif text == "start":
            columns.setdefault("start", cell.column)
        elif text == "end":
            columns.setdefault("end", cell.column)
        elif text == "pay code":
            columns.setdefault("pay_code", cell.column)
        elif text == "shift label":
            columns.setdefault("shift_label", cell.column)

    required = {"job", "employee", "start", "end"}
    if not required.issubset(columns):
        return None

    kind = "pay" if "pay_code" in columns else "schedule"
    return HeaderMap(
        kind=kind,
        job_col=columns["job"],
        employee_col=columns["employee"],
        start_col=columns["start"],
        end_col=columns["end"],
        pay_code_col=columns.get("pay_code"),
        shift_label_col=columns.get("shift_label"),
    )


def _parse_schedule_row(row: tuple[Any, ...], header: HeaderMap, sheet_name: str) -> ScheduleRecord | None:
    job_raw = _as_text(_cell_value(row, header.job_col))
    employee = _as_text(_cell_value(row, header.employee_col))
    start_value = _cell_value(row, header.start_col)
    end_value = _cell_value(row, header.end_col)

    if not job_raw or not employee:
        return None
    if _normalize_header(job_raw) == "job" or _normalize_header(employee) == "employee":
        return None
    if not isinstance(start_value, datetime) or not isinstance(end_value, datetime):
        return None

    job_code = extract_job_code(job_raw)
    if is_excluded_job(job_code):
        return None

    work_date, shift_kind = classify_shift(start_value)
    role_group = classify_role(job_code)

    return ScheduleRecord(
        employee_display_name=employee,
        work_date=work_date,
        shift_kind=shift_kind,
        start_time=start_value.time().replace(second=0, microsecond=0),
        end_time=end_value.time().replace(second=0, microsecond=0),
        job_code=job_code,
        role_group=role_group,
        is_clinical_leader=job_code == CLINICAL_LEADER_CODE,
        source_evidence=[
            f"{sheet_name}!{get_column_letter(header.job_col)}{row[0].row}: schedule job",
            f"{sheet_name}!{get_column_letter(header.start_col)}{row[0].row}: schedule start",
            f"{sheet_name}!{get_column_letter(header.end_col)}{row[0].row}: schedule end",
        ],
        source_sheet=sheet_name,
        source_row=row[0].row,
    )


def _parse_pay_lead_row(row: tuple[Any, ...], header: HeaderMap, sheet_name: str) -> PayLeadEvidence | None:
    if header.pay_code_col is None:
        return None

    pay_code_text = _as_text(_cell_value(row, header.pay_code_col))
    if not is_lead_pay(pay_code_text):
        return None

    job_raw = _as_text(_cell_value(row, header.job_col))
    employee = _as_text(_cell_value(row, header.employee_col))
    if not job_raw or not employee:
        return None

    start_time = _coerce_time(_cell_value(row, header.start_col))
    end_time = _coerce_time(_cell_value(row, header.end_col))
    job_code = extract_job_code(job_raw)
    if is_excluded_job(job_code):
        return None

    source_row = row[0].row

    return PayLeadEvidence(
        employee_display_name=employee,
        job_code_raw=job_raw,
        job_code=job_code,
        start_time=start_time,
        end_time=end_time,
        source_sheet=sheet_name,
        source_row=source_row,
        source_evidence=[
            f"{sheet_name}!{get_column_letter(header.pay_code_col)}{source_row}: lead pay code",
        ],
    )


def _attach_pay_leads(
    records: list[ScheduleRecord],
    pay_leads: list[PayLeadEvidence],
    warnings: list[str],
) -> None:
    for pay_lead in pay_leads:
        shift_kind = classify_shift_from_time(pay_lead.start_time)
        context = _nearest_context_record(records, pay_lead, shift_kind)
        work_date = context.work_date if context else None

        if not work_date and pay_lead.warning:
            warnings.append(pay_lead.warning)
            continue
        if not work_date:
            warnings.append("Found a lead-pay row without enough nearby schedule context.")
            continue

        matched = _find_matching_record(records, pay_lead, work_date, shift_kind)
        if matched:
            matched.has_lead_pay = True
            matched.source_evidence.extend(pay_lead.source_evidence)
            continue

        start_time = pay_lead.start_time
        end_time = pay_lead.end_time
        if not start_time or not end_time:
            warnings.append("Found a lead-pay row without readable start or end time.")
            continue

        inferred_warning = "Lead-pay-only row was included using nearby schedule context."
        records.append(
            ScheduleRecord(
                employee_display_name=pay_lead.employee_display_name,
                work_date=work_date,
                shift_kind=shift_kind,
                start_time=start_time,
                end_time=end_time,
                job_code=pay_lead.job_code,
                role_group=classify_role(pay_lead.job_code),
                is_clinical_leader=pay_lead.job_code == CLINICAL_LEADER_CODE,
                has_lead_pay=True,
                source_evidence=pay_lead.source_evidence
                + [f"{context.source_sheet}!row {context.source_row}: inferred date/shift context"],
                warnings=[inferred_warning],
                source_sheet=pay_lead.source_sheet,
                source_row=pay_lead.source_row,
            )
        )


def _nearest_context_record(
    records: list[ScheduleRecord],
    pay_lead: PayLeadEvidence,
    shift_kind: str,
) -> ScheduleRecord | None:
    candidates = [
        record
        for record in records
        if record.source_sheet == pay_lead.source_sheet
        and record.source_row is not None
        and record.source_row < pay_lead.source_row
        and record.shift_kind == shift_kind
        and record.job_code == pay_lead.job_code
        and pay_lead.source_row - record.source_row <= 60
    ]
    if not candidates:
        candidates = [
            record
            for record in records
            if record.source_sheet == pay_lead.source_sheet
            and record.source_row is not None
            and record.source_row < pay_lead.source_row
            and record.shift_kind == shift_kind
            and pay_lead.source_row - record.source_row <= 60
        ]
    return max(candidates, key=lambda record: record.source_row or 0) if candidates else None


def _find_matching_record(
    records: list[ScheduleRecord],
    pay_lead: PayLeadEvidence,
    work_date: date,
    shift_kind: str,
) -> ScheduleRecord | None:
    employee_key = _employee_key(pay_lead.employee_display_name)
    matches = [
        record
        for record in records
        if _employee_key(record.employee_display_name) == employee_key
        and record.work_date == work_date
        and record.shift_kind == shift_kind
    ]
    if not matches:
        return None
    same_code = [record for record in matches if record.job_code == pay_lead.job_code]
    return same_code[0] if same_code else matches[0]


def extract_job_code(job_raw: str) -> str:
    text = str(job_raw).strip()
    for code in sorted(KNOWN_JOB_CODES, key=len, reverse=True):
        if re.search(rf"(?<!\w){re.escape(code)}(?!\w)", text):
            return code
    match = re.match(r"([A-Za-z0-9]+)", text)
    return match.group(1).upper() if match else text.upper()


def classify_role(job_code: str) -> str:
    if job_code in RN_CODES:
        return ROLE_RN
    if job_code in PSA_CODES:
        return ROLE_PSA
    return ROLE_OTHER


def is_excluded_job(job_code: str) -> bool:
    return job_code in EXCLUDED_JOB_CODES


def classify_shift(start_datetime: datetime) -> tuple[date, str]:
    start = start_datetime.time()
    if start.hour >= 15:
        return start_datetime.date(), SHIFT_NIGHT
    if start.hour < 7:
        return start_datetime.date() - timedelta(days=1), SHIFT_NIGHT
    return start_datetime.date(), SHIFT_DAY


def classify_shift_from_time(value: time | None) -> str:
    if value is None:
        return SHIFT_DAY
    if value.hour >= 15 or value.hour < 7:
        return SHIFT_NIGHT
    return SHIFT_DAY


def is_lead_pay(pay_code: str | None) -> bool:
    if not pay_code:
        return False
    normalized = _normalize_pay_code(pay_code)
    return normalized in LEAD_PAY_CODES


def _normalize_header(value: Any) -> str:
    text = _as_text(value)
    if not text:
        return ""
    normalized = re.sub(r"\s+", " ", text.strip().lower())
    if normalized in {"job", "job code"}:
        return "job"
    if normalized in {"employee", "employee name", "person", "name"}:
        return "employee"
    if normalized in {"start", "start time", "shift start"}:
        return "start"
    if normalized in {"end", "end time", "shift end"}:
        return "end"
    if normalized in {"pay code", "pay codes"}:
        return "pay code"
    if normalized in {"shift label", "shift"}:
        return "shift label"
    return normalized


def _normalize_pay_code(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().upper())


def _employee_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().casefold())


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_value(row: tuple[Any, ...], one_based_col: int) -> Any:
    index = one_based_col - 1
    if index < 0 or index >= len(row):
        return None
    return row[index].value


def _coerce_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    return None
