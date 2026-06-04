"""Build synthetic MyTime-like source workbooks for tests."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook


def build_source_workbook(rows: list[dict]) -> bytes:
    """Create a synthetic workbook with repeated schedule and pay-code sections."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    current_row = 1
    for section in rows:
        current_row = _write_schedule_section(sheet, current_row, section.get("schedule", []))
        if section.get("pay_codes"):
            current_row = _write_pay_section(sheet, current_row, section["pay_codes"])
        current_row += 2

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_schedule_section(sheet, row: int, schedule_rows: list[dict]) -> int:
    sheet.cell(row, 1, "Job")
    sheet.cell(row, 5, "Employee")
    sheet.cell(row, 18, "Start")
    sheet.cell(row, 21, "End")
    sheet.cell(row, 25, "Shift Label")
    row += 1

    for item in schedule_rows:
        sheet.cell(row, 1, item["job"])
        sheet.cell(row, 5, item["employee"])
        sheet.cell(row, 18, item["start"])
        sheet.cell(row, 21, item["end"])
        sheet.cell(row, 25, item.get("shift_label", ""))
        row += 1

    return row


def _write_pay_section(sheet, row: int, pay_rows: list[dict]) -> int:
    sheet.cell(row, 2, "Job")
    sheet.cell(row, 7, "Employee")
    sheet.cell(row, 14, "Pay Code")
    sheet.cell(row, 19, "Start")
    sheet.cell(row, 23, "End")
    row += 1

    for item in pay_rows:
        sheet.cell(row, 2, item["job"])
        sheet.cell(row, 7, item["employee"])
        sheet.cell(row, 14, item["pay_code"])
        sheet.cell(row, 19, item.get("start", datetime(1970, 1, 1, 19, 0)))
        sheet.cell(row, 23, item.get("end", datetime(1970, 1, 1, 7, 0)))
        row += 1

    return row

