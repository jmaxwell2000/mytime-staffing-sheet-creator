from datetime import date, time
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from app.parser.constants import ROLE_PSA, ROLE_RN, SHIFT_DAY, SHIFT_NIGHT
from app.parser.models import ScheduleRecord
from app.workbook.build_workbook import (
    BACK_FOOTER_FONT_SIZE,
    BACK_FOOTER_TEXT,
    BACK_LINE_ROWS,
    BACK_PENCIL_REMINDER,
    COLUMN_WIDTHS,
    FALL_PREVENTION_TEXT,
    FRONT_PENCIL_REMINDER,
    BOTTOM_TITLE_SPACER_HEIGHT,
    PSA_CAPACITY,
    RN_CAPACITY,
    STAFF_NAME_BASE_FONT_SIZE,
    STAFF_NAME_FIT_CHARS_AT_BASE_SIZE,
    STAFF_NAME_MIN_FONT_SIZE,
    STAFF_ROW_HEIGHT,
    STAFF_TABLE_FONT_SIZE,
    build_staffing_workbook,
    resolve_lead_display,
)

EXPECTED_FALL_PREVENTION_TEXT = (
    "Post-fall measures (circle):         Alarm          Nonskid socks          Fall sign"
    "          Stay with pt toileting          Fall risk score          Other                  "
)


def record(
    name: str,
    job_code: str,
    *,
    role_group: str = ROLE_RN,
    work_date: date = date(2026, 4, 28),
    shift_kind: str = SHIFT_NIGHT,
    start: time = time(19, 0),
    end: time = time(7, 30),
    clinical_leader: bool = False,
    lead_pay: bool = False,
) -> ScheduleRecord:
    return ScheduleRecord(
        employee_display_name=name,
        work_date=work_date,
        shift_kind=shift_kind,
        start_time=start,
        end_time=end,
        job_code=job_code,
        role_group=role_group,
        is_clinical_leader=clinical_leader,
        has_lead_pay=lead_pay,
    )


def test_clinical_leader_overrides_lead_pay_regression():
    records = [
        record("Employee 001", "3178", lead_pay=True),
        record("Employee 002", "8536", clinical_leader=True),
    ]

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    lead_cells = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.value == "LEAD"]
    assert lead_cells == ["F3"]
    assert sheet["C3"].value == "Employee 002"
    assert sheet["C4"].value == "Employee 001"
    assert sheet["F3"].alignment.horizontal == "center"


def test_names_are_normalized_and_sorted_with_lead_rn_first():
    records = [
        record("bETAUSER, bRAVO", "3178"),
        record("zETAUSER, aLPHA", "8536", clinical_leader=True),
        record("aLPHAUSER, cHARLIE", "3178"),
        record("pSAUSER, dELTA", "5056", role_group=ROLE_PSA),
    ]

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    assert sheet["C3"].value == "Zetauser, Alpha"
    assert sheet["F3"].value == "LEAD"
    assert sheet["C4"].value == "Alphauser, Charlie"
    assert sheet["C5"].value == "Betauser, Bravo"
    assert sheet["C20"].value == "Psauser, Delta"


def test_multiple_clinical_leaders_get_cl_suffix_and_no_lead():
    records = [
        record("Employee 001", "8536", clinical_leader=True),
        record("Employee 002", "8536", clinical_leader=True),
    ]

    display = resolve_lead_display(records)

    assert display[id(records[0])] == ("Employee 001 (CL)", False)
    assert display[id(records[1])] == ("Employee 002 (CL)", False)


def test_workbook_builder_excludes_staff_lpn_records_defensively():
    records = [
        record("Employee 001", "3178"),
        record("Employee 002", "4120"),
    ]

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    visible_values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
    assert "Employee 001" in visible_values
    assert "Employee 002" not in visible_values


def test_staff_name_font_size_uses_final_cl_display_text():
    records = [
        record("aBCDEFGHIJKLMNOPQRSTUVWXYZ, kLMNOPQRSTUVWX", "8536", clinical_leader=True),
        record("pQRSTUVWXYZABCDEFGHIJKLM, qRSTUVWXYZABC", "8536", clinical_leader=True),
    ]

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    assert sheet["C3"].value == "Abcdefghijklmnopqrstuvwxyz, Klmnopqrstuvwx (CL)"
    assert sheet["C3"].font.sz == _expected_name_font_size(sheet["C3"].value)
    assert sheet["C3"].font.sz < STAFF_NAME_BASE_FONT_SIZE
    assert sheet["C3"].font.sz >= STAFF_NAME_MIN_FONT_SIZE
    assert sheet["C3"].alignment.shrinkToFit is True
    assert sheet["C4"].value == "Pqrstuvwxyzabcdefghijklm, Qrstuvwxyzabc (CL)"
    assert sheet["C4"].font.sz == _expected_name_font_size(sheet["C4"].value)
    assert sheet["C4"].font.sz < STAFF_NAME_BASE_FONT_SIZE
    assert sheet["C4"].font.sz >= STAFF_NAME_MIN_FONT_SIZE
    assert sheet["C4"].alignment.shrinkToFit is True


def test_single_cl_name_stays_normal_size_when_displayed_as_lead():
    records = [
        record("aBCDEFGHIJ, kLMNO", "8536", clinical_leader=True),
        record("pQRSTUVWXYZ, qRSTU", "3178"),
    ]

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    assert sheet["C3"].value == "Abcdefghij, Klmno"
    assert sheet["C3"].font.sz == STAFF_NAME_BASE_FONT_SIZE
    assert sheet["F3"].value == "LEAD"


def test_workbook_front_and_back_structure_and_time_styling():
    records = [
        record(
            "Employee 001",
            "3178",
            work_date=date(2026, 4, 29),
            shift_kind=SHIFT_DAY,
            start=time(11, 0),
            end=time(15, 0),
        ),
        record(
            "Employee 002",
            "5056",
            role_group=ROLE_PSA,
            work_date=date(2026, 4, 29),
            shift_kind=SHIFT_DAY,
            start=time(7, 0),
            end=time(19, 0),
        ),
    ]

    output = build_staffing_workbook(records, include_back_pages=True, back_page_offset=0.05)
    workbook = load_workbook(BytesIO(output))
    rich_workbook = load_workbook(BytesIO(output), rich_text=True)

    assert workbook.sheetnames == ["Wed Apr29 DAY", "Wed Apr29 DAY BACK"]
    front = workbook["Wed Apr29 DAY"]
    back = workbook["Wed Apr29 DAY BACK"]
    rich_front = rich_workbook["Wed Apr29 DAY"]
    rich_back = rich_workbook["Wed Apr29 DAY BACK"]
    assert str(front.print_area) == "'Wed Apr29 DAY'!$B$1:$H$39"
    assert front.page_setup.orientation == "portrait"
    assert int(front.page_setup.paperSize) == int(front.PAPERSIZE_LETTER)
    assert front.page_setup.fitToWidth == 1
    assert front.page_setup.fitToHeight == 1
    assert front.sheet_properties.pageSetUpPr.fitToPage is True
    assert front.page_margins.left == 0.75
    assert front.page_margins.right == 0.25
    assert str(front.print_area) == "'Wed Apr29 DAY'!$B$1:$H$39"
    assert front.column_dimensions["B"].width == COLUMN_WIDTHS["B"]
    assert front.column_dimensions["B"].width == 5.01
    assert front.column_dimensions["C"].width == 35.2
    assert front.column_dimensions["C"].width == COLUMN_WIDTHS["C"]
    assert front.column_dimensions["D"].width == 7.0
    assert front.column_dimensions["E"].width == 7.0
    for column in "FGH":
        assert front.column_dimensions[column].width == 29.16
        assert front.column_dimensions[column].width == COLUMN_WIDTHS[column]
    for column in "BCDEFGH":
        assert back.column_dimensions[column].width == front.column_dimensions[column].width
    assert round(sum(front.column_dimensions[column].width for column in "BCDEFGH"), 2) == 141.69
    assert front["B1"].value == FRONT_PENCIL_REMINDER
    assert front["B1"].alignment.horizontal == "left"
    assert front["G1"].value == "Wednesday Day Shift \u25ef 4/29"
    assert front["G1"].alignment.horizontal == "right"
    assert back["B1"].value == "4/29 \u25ef Wednesday Day Shift"
    assert back["B1"].alignment.horizontal == "left"
    assert back["G1"].value == BACK_PENCIL_REMINDER
    assert back["G1"].alignment.horizontal == "right"
    assert _rich_parts(rich_front["G1"]) == [
        ("Wednesday Day Shift ", 16.0, False, None),
        ("\u25ef", 22.0, True, None),
        (" 4/29", 22.0, True, None),
    ]
    assert _rich_parts(rich_front["B39"]) == [
        ("Wednesday Day Shift ", 16.0, False, None),
        ("\u25ef", 22.0, True, None),
        (" 4/29", 22.0, True, None),
    ]
    assert _rich_parts(rich_back["B1"]) == [
        ("4/29 ", 22.0, True, "A6A6A6"),
        ("\u25ef", 22.0, True, "A6A6A6"),
        (" Wednesday Day Shift", 16.0, False, "A6A6A6"),
    ]
    assert _rich_parts(rich_back["B39"]) == [
        ("4/29 ", 22.0, True, "A6A6A6"),
        ("\u25ef", 22.0, True, "A6A6A6"),
        (" Wednesday Day Shift", 16.0, False, "A6A6A6"),
    ]
    assert front["B2"].value == "Initial"
    assert front["B2"].font.sz == 7
    assert front["B2"].alignment.shrinkToFit is True
    assert front["D2"].font.sz == STAFF_TABLE_FONT_SIZE
    assert front["E2"].font.sz == STAFF_TABLE_FONT_SIZE
    assert front["C19"].value == "PSA"
    assert front["B19"].value == "Initial"
    assert front.row_dimensions[3].height == STAFF_ROW_HEIGHT
    assert front.row_dimensions[20].height == STAFF_ROW_HEIGHT
    assert front["B31"].value.startswith("Patient Fall Review (Wed 4/29) - Charge RN:")
    assert front["B31"].font.bold is True
    assert front["B31"].font.sz == 10
    assert front["B31"].alignment.horizontal == "center"
    assert front.row_dimensions[31].height == 26
    assert front.row_dimensions[33].height == 30
    assert front.row_dimensions[37].height == 30
    assert "Yes/No" not in front["B31"].value
    assert "answer the questions" in front["B31"].value
    assert front["B32"].value == "Question"
    assert front["B32"].font.sz == 10
    assert front["E32"].value is None
    assert front["F32"].value == "Answer / Notes"
    assert front["G32"].value is None
    assert front["F32"].font.sz == 10
    assert front["B33"].value == "\u2022 Bathroom or bedside commode?"
    assert front["B33"].font.sz == 10
    assert front["B33"].alignment.indent == 1
    assert front["F33"].value is None
    assert front["F33"].border.bottom.style == "thin"
    assert front["B34"].value == "\u2022 Staff or family with patient at time of fall?"
    assert front["B35"].value == "\u2022 If no, should someone have been present?"
    assert front["B36"].value == "\u2022 Any injury?"
    assert FALL_PREVENTION_TEXT == EXPECTED_FALL_PREVENTION_TEXT
    assert front["B37"].value == EXPECTED_FALL_PREVENTION_TEXT
    assert "\n" not in front["B37"].value
    assert "|" not in front["B37"].value
    assert "_" not in front["B37"].value
    assert front["B37"].font.sz == 10
    assert front["B37"].font.bold is False
    assert front["B37"].alignment.horizontal == "left"
    assert front["B37"].alignment.wrapText is not True
    assert front["B37"].alignment.shrinkToFit is not True
    assert front.row_dimensions[38].height == BOTTOM_TITLE_SPACER_HEIGHT
    assert front["D3"].value == 1100
    assert front["D3"].number_format == "0000"
    assert front["D3"].font.bold is False
    assert front["D3"].font.italic is False
    assert front["D3"].font.sz == STAFF_TABLE_FONT_SIZE
    assert _rgb(front["D3"].fill.fgColor) == "FFF2CC"
    assert front["E3"].value == 1500
    assert front["E3"].number_format == "0000"
    assert front["E3"].font.bold is False
    assert front["E3"].font.italic is False
    assert front["E3"].font.sz == STAFF_TABLE_FONT_SIZE
    assert _rgb(front["E3"].fill.fgColor) == "FFF2CC"
    assert front["F3"].font.sz == STAFF_TABLE_FONT_SIZE
    assert front["D20"].font.sz == STAFF_TABLE_FONT_SIZE
    assert front["E20"].value == 1930
    assert front["E20"].number_format == "0000"
    assert int(back.page_setup.paperSize) == int(back.PAPERSIZE_LETTER)
    assert back.page_setup.fitToWidth == 1
    assert back.page_setup.fitToHeight == 1
    assert back.sheet_properties.pageSetUpPr.fitToPage is True
    assert back.page_margins.left == 0.25
    assert back.page_margins.right == 0.75
    assert back.page_margins.top == 0.2
    for row in range(1, 40):
        assert back.row_dimensions[row].height == front.row_dimensions[row].height

    for row in range(2, 38):
        for column in range(2, 9):
            border = front.cell(row, column).border
            for side in (border.left, border.right, border.top, border.bottom):
                assert side.style in (None, "thin")

    front_grid_rows = [
        row
        for row in range(2, 31)
        if any(front.cell(row, column).border.bottom.style for column in range(2, 9))
    ]
    back_line_rows = [
        row
        for row in range(2, 31)
        if any(back.cell(row, column).border.bottom.style for column in range(2, 9))
    ]
    assert tuple(back_line_rows) == BACK_LINE_ROWS
    assert 2 not in back_line_rows
    assert all(row in front_grid_rows for row in back_line_rows)
    assert front["G1"].alignment.horizontal == "right"
    assert front["B39"].alignment.horizontal == "right"
    assert front["B39"].value == "Wednesday Day Shift \u25ef 4/29"
    assert back["B1"].alignment.horizontal == "left"
    assert back["B39"].alignment.horizontal == "left"
    assert back["B39"].value == "4/29 \u25ef Wednesday Day Shift"
    assert back["B31"].value is None
    assert back["B38"].value == BACK_FOOTER_TEXT
    assert back["B38"].font.sz == BACK_FOOTER_FONT_SIZE
    assert back["B38"].alignment.horizontal == "center"
    assert _rgb(back["B1"].font.color) == "A6A6A6"
    assert _rgb(back["G1"].font.color) == "A6A6A6"
    assert _rgb(back["B3"].font.color) == "A6A6A6"
    assert _rgb(back["B38"].font.color) == "A6A6A6"
    assert back["B3"].border.bottom.style == "thin"
    assert _rgb(back["B3"].border.bottom.color) == "D9D9D9"


def test_full_rn_and_psa_sections_keep_one_blank_writable_row():
    records = [
        record(f"Rn {index:03d}", "3178", start=time(19, 0), end=time(7, 30))
        for index in range(RN_CAPACITY)
    ]
    records.extend(
        record(f"Psa {index:03d}", "5056", role_group=ROLE_PSA, start=time(19, 0), end=time(7, 30))
        for index in range(PSA_CAPACITY)
    )

    output = build_staffing_workbook(records)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Tue Apr28 NIGHT"]

    assert sheet["C19"].value is None
    assert sheet["C19"].border.bottom.style == "thin"
    assert sheet["C20"].value == "PSA"
    assert sheet["C32"].value is None
    assert sheet["C32"].border.bottom.style == "thin"
    assert sheet["B33"].value.startswith("Patient Fall Review (Tue 4/28) - Charge RN:")
    assert str(sheet.print_area) == "'Tue Apr28 NIGHT'!$B$1:$H$41"


def _rgb(color) -> str | None:
    if isinstance(color, str):
        return color[-6:]
    if color is None or color.type != "rgb" or color.rgb is None:
        return None
    return color.rgb[-6:]


def _rich_parts(cell) -> list[tuple[str, float, bool, str | None]]:
    value = cell.value
    assert isinstance(value, CellRichText)
    return [(part.text, float(part.font.sz), bool(part.font.b), _rgb(part.font.color)) for part in value]


def _expected_name_font_size(display_name: str) -> int:
    if len(display_name) <= STAFF_NAME_FIT_CHARS_AT_BASE_SIZE:
        return STAFF_NAME_BASE_FONT_SIZE
    fitted_size = int(STAFF_NAME_BASE_FONT_SIZE * STAFF_NAME_FIT_CHARS_AT_BASE_SIZE / len(display_name))
    return max(STAFF_NAME_MIN_FONT_SIZE, fitted_size)
