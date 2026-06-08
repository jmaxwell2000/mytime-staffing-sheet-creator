from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import __updated_date__, __version__
from app.main import app
from tests.fixtures.source_workbooks import build_source_workbook


client = TestClient(app)


def test_healthz():
    response = client.get("/healthz")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_index_has_stateful_upload_flow_and_visible_offset_control():
    response = client.get("/")
    assert response.status_code == 200
    html = response.text
    assert app.title == "MyTime Staffing Sheet Creator"
    assert __version__ == "2.1"
    assert __updated_date__ == "June 8, 2026"
    assert "MyTime Staffing Sheet Creator" in html
    assert "v2.1" in html
    assert "Updated June 8, 2026" in html
    assert "{{ app_version }}" not in html
    assert "{{ updated_date }}" not in html
    assert "MyTime 3.0" not in html
    assert 'method="post"' in html
    assert 'action="/convert"' in html
    assert 'id="options-panel" class="options"' in html
    assert 'id="offset-control" class="offset-control"' in html
    assert 'id="offset-control" class="offset-control" hidden' not in html
    assert 'id="submit-button-label">Upload file</span>' in html
    assert "viewed through paper" in html


def test_static_js_uses_stateful_fetch_download_and_persisted_defaults():
    response = client.get(f"/static/app.js?v={__version__}")
    assert response.status_code == 200
    script = response.text
    assert "event.preventDefault" in script
    assert "fetch(form.action" in script
    assert "Content-Disposition" in script
    assert "URL.createObjectURL" in script
    assert "URL.revokeObjectURL" in script
    assert 'mytime.includeBackPages' in script
    assert 'mytime.backPageOffset' in script
    assert 'Upload file' in script
    assert 'Create staffing sheets' in script
    assert 'Creating...' in script
    assert 'Downloaded' in script


def test_static_css_uses_blue_theme_without_state_button_colors():
    response = client.get(f"/static/app.css?v={__version__}")
    assert response.status_code == 200
    stylesheet = response.text
    assert "--accent: #2563eb;" in stylesheet
    assert "--accent-dark: #1d4ed8;" in stylesheet
    assert "--accent-soft: #eaf1ff;" in stylesheet
    assert "--accent-glow: rgba(37, 99, 235, 0.16);" in stylesheet
    assert 'primary-button[data-state="processing"]' in stylesheet
    assert "animation: buttonSweep" in stylesheet
    assert "--processing" not in stylesheet
    assert "--success" not in stylesheet


def test_release_metadata_is_consistent_across_public_docs():
    expected = f"Current release: v{__version__}. Updated {__updated_date__}."
    for path in ("README.md", "docs/product-brief.md", "docs/requirements.md"):
        content = open(path, encoding="utf-8").read()
        assert expected in content


def test_convert_returns_xlsx_download(tmp_path):
    data = build_source_workbook(
        [
            {
                "schedule": [
                    {
                        "job": "8536-RN Clinical Leader",
                        "employee": "Employee 001",
                        "start": datetime(2026, 4, 28, 19, 0),
                        "end": datetime(2026, 4, 29, 7, 30),
                    }
                ]
            }
        ]
    )

    response = client.post(
        "/convert",
        files={"workbook": ("source.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"include_back_pages": "false", "back_page_offset": "0.00"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="staffing-sheets-2026-04-28.xlsx"' in response.headers["content-disposition"]

    output = tmp_path / "output.xlsx"
    output.write_bytes(response.content)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Tue Apr28 NIGHT"]
    assert workbook["Tue Apr28 NIGHT"]["F3"].value == "LEAD"


def test_convert_rejects_non_xlsx():
    response = client.post(
        "/convert",
        files={"workbook": ("source.txt", b"not a workbook", "text/plain")},
        data={"include_back_pages": "false", "back_page_offset": "0.00"},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Upload a valid .xlsx workbook."


def test_convert_rejects_bad_offset():
    data = build_source_workbook(
        [
            {
                "schedule": [
                    {
                        "job": "3178-Staff RN Placeholder",
                        "employee": "Employee 001",
                        "start": datetime(2026, 4, 28, 7, 0),
                        "end": datetime(2026, 4, 28, 19, 30),
                    }
                ]
            }
        ]
    )

    response = client.post(
        "/convert",
        files={"workbook": ("source.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"include_back_pages": "true", "back_page_offset": "0.50"},
    )

    assert response.status_code == 400
    assert "Back page offset" in response.json()["detail"]
